from fastapi import FastAPI, Request, Form
from fastapi.responses import HTMLResponse, RedirectResponse
from fastapi.templating import Jinja2Templates
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import os
import re
from datetime import datetime
from threading import Lock

APP_DIR = os.path.dirname(os.path.abspath(__file__))
XLSX_PATH = os.path.join(APP_DIR, "customers.xlsx")
SHEET_NAME = "Customers"
templates = Jinja2Templates(directory=os.path.join(APP_DIR, "templates"))
file_lock = Lock()

HEADERS = [
    "CustomerID", "利用者名称", "利用開始日", "介護区分", "連絡先（電話）", "住所", "メールアドレス",
    "担当者", "注意事項", "タイムスタンプ", "データ編集リンク"
]

PHONE_RE = re.compile(r"^0\d{1,4}-?\d{1,4}-?\d{4}$")
EMAIL_RE = re.compile(r"^[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}$")

app = FastAPI()


def ensure_workbook():
    if os.path.exists(XLSX_PATH):
        wb = load_workbook(XLSX_PATH)
        if SHEET_NAME not in wb.sheetnames:
            ws = wb.create_sheet(SHEET_NAME)
            ws.append(HEADERS)
        return wb

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    ws.append(HEADERS)

    # 列幅とかはOK（これはmax_rowに影響しない）
    ws.column_dimensions["A"].width = 12

    wb.save(XLSX_PATH)
    return wb


def next_customer_id(ws):
    # A列（CustomerID）最大値+1（4桁）
    max_num = 0
    for r in range(2, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if v is None or v == "":
            continue
        try:
            n = int(str(v))
            if n > max_num:
                max_num = n
        except ValueError:
            pass
    return str(max_num + 1).zfill(4)


def get_all_rows(ws):
    rows = []
    for r in range(2, ws.max_row + 1):
        cid = ws.cell(r, 1).value
        if cid is None or cid == "":
            continue
        row = {HEADERS[c - 1]: ws.cell(r, c).value for c in range(1, len(HEADERS) + 1)}
        rows.append(row)
    return rows


def find_row_by_customer_id(ws, customer_id: str):
    for r in range(2, ws.max_row + 1):
        if str(ws.cell(r, 1).value).strip() == customer_id:
            return r
    return None


@app.get("/", response_class=HTMLResponse)
def root():
    return RedirectResponse("/customers")


@app.get("/customers", response_class=HTMLResponse)
def list_customers(request: Request):
    with file_lock:
        wb = ensure_workbook()
        ws = wb[SHEET_NAME]
        rows = get_all_rows(ws)
        wb.save(XLSX_PATH)
        print("SAVED:", XLSX_PATH)
    return templates.TemplateResponse("list.html", {"request": request, "rows": rows})


@app.get("/customers/new", response_class=HTMLResponse)
def new_customer(request: Request):
    with file_lock:
        wb = ensure_workbook()
        ws = wb[SHEET_NAME]
        cid = next_customer_id(ws)
        wb.save(XLSX_PATH)
        print("SAVED:", XLSX_PATH)
    return templates.TemplateResponse("form.html", {"request": request, "customer_id": cid, "error": None})


@app.post("/customers", response_class=HTMLResponse)
def create_customer(
    request: Request,
    customer_id: str = Form(...),
    name: str = Form(...),
    start_date: str = Form(""),
    care: str = Form(""),
    phone: str = Form(""),
    address: str = Form(""),
    email: str = Form(""),
    staff: str = Form(""),
    note: str = Form(""),
):
    # 簡易バリデーション
    if phone and not PHONE_RE.match(phone):
        return templates.TemplateResponse("form.html", {"request": request, "customer_id": customer_id, "error": "電話番号の形式が不正です（例：090-1234-5678）"})
    if email and not EMAIL_RE.match(email):
        return templates.TemplateResponse("form.html", {"request": request, "customer_id": customer_id, "error": "メールアドレスの形式が不正です（例：name@example.com）"})

    timestamp = datetime.now().strftime("%Y/%m/%d %H:%M:%S")
    edit_link = f"/customers/{customer_id}/edit"

    with file_lock:
        wb = ensure_workbook()
        ws = wb[SHEET_NAME]

        # 既に同じCustomerIDがあれば更新扱いにする（欠番対策にもなる）
        rownum = find_row_by_customer_id(ws, customer_id)
        values = [customer_id, name, start_date, care, phone, address, email, staff, note, timestamp, edit_link]

        if rownum:
            for c, v in enumerate(values, start=1):
                ws.cell(rownum, c).value = v
            ws[f"A{rownum}"].number_format = "@"
        else:
            ws.append(values)
            ws[f"A{ws.max_row}"].number_format = "@"

        wb.save(XLSX_PATH)
        print("SAVED:", XLSX_PATH)
    return RedirectResponse("/customers", status_code=303)


@app.get("/customers/{customer_id}/edit", response_class=HTMLResponse)
def edit_customer(request: Request, customer_id: str):
    with file_lock:
        wb = ensure_workbook()
        ws = wb[SHEET_NAME]
        rownum = find_row_by_customer_id(ws, customer_id)
        if not rownum:
            wb.save(XLSX_PATH)
            print("SAVED:", XLSX_PATH)
            return RedirectResponse("/customers", status_code=303)

        data = {HEADERS[c - 1]: ws.cell(rownum, c).value for c in range(1, len(HEADERS) + 1)}
        wb.save(XLSX_PATH)
        print("SAVED:", XLSX_PATH)
    return templates.TemplateResponse("edit.html", {"request": request, "data": data, "error": None})


@app.post("/customers/{customer_id}/edit")
def update_customer(
    request: Request,
    customer_id: str,
    name: str = Form(...),
    start_date: str = Form(""),
    care: str = Form(""),
    phone: str = Form(""),
    address: str = Form(""),
    email: str = Form(""),
    staff: str = Form(""),
    note: str = Form(""),
):
    if phone and not PHONE_RE.match(phone):
        return templates.TemplateResponse("edit.html", {"request": request, "data": {"CustomerID": customer_id, "利用者名称": name, "利用開始日": start_date, "介護区分": care, "連絡先（電話）": phone, "住所": address, "メールアドレス": email, "担当者": staff, "注意事項": note}, "error": "電話番号の形式が不正です"})
    if email and not EMAIL_RE.match(email):
        return templates.TemplateResponse("edit.html", {"request": request, "data": {"CustomerID": customer_id, "利用者名称": name, "利用開始日": start_date, "介護区分": care, "連絡先（電話）": phone, "住所": address, "メールアドレス": email, "担当者": staff, "注意事項": note}, "error": "メールアドレスの形式が不正です"})

    timestamp = datetime.now().strftime("%Y/%m/%d %H:%M:%S")
    edit_link = f"/customers/{customer_id}/edit"

    with file_lock:
        wb = ensure_workbook()
        ws = wb[SHEET_NAME]
        rownum = find_row_by_customer_id(ws, customer_id)
        if rownum:
            values = [customer_id, name, start_date, care, phone, address, email, staff, note, timestamp, edit_link]
            for c, v in enumerate(values, start=1):
                ws.cell(rownum, c).value = v
            ws[f"A{rownum}"].number_format = "@"
            wb.save(XLSX_PATH)
            print("SAVED:", XLSX_PATH)
    return RedirectResponse("/customers", status_code=303)
